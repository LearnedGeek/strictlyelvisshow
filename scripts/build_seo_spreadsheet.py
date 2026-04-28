import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

INPUT = "e:/Documents/Work/strictlyelvisshow/docs/client/SJOA_Page_Titles_and_Descriptions_FULL.xlsx"
OUTPUT = "e:/Documents/Work/strictlyelvisshow/docs/client/SJOA_Page_Titles_and_Descriptions_PROPOSED.xlsx"

wb = openpyxl.load_workbook(INPUT)
ws = wb["Sheet1"]

proposals = {
    "/": (
        "St. Joan of Arc Catholic K3-8 | Nashotah, WI",
        "Faith-centered Catholic K3-8 school serving Lake Country families since 1847. Multi-age classrooms, K3/K4 programs, after-school care. Schedule a tour."
    ),
    "/about-us": (
        "About Our Catholic School | Lake Country, WI",
        "Learn about St. Joan of Arc - a JK-8 Catholic school in Lake Country, WI. WRISA accredited, certified teachers, and a nurturing faith-based community."
    ),
    "/school-history": (
        "Our History Since 1847 | SJOA Nashotah, WI",
        "From our founding in Okauchee to growth in Nashotah, discover the 175+ year legacy of St. Joan of Arc Parish School - serving Lake Country families."
    ),
    "/our-foundation": (
        "Our Foundation | Faith & Catholic Values | SJOA",
        "The foundation of St. Joan of Arc nurtures faith, fellowship, and family. Shaping compassionate future leaders through Catholic education in Nashotah, WI."
    ),
    "/academics": (
        "Academics | Catholic K3-8 Curriculum | SJOA",
        "Explore our accredited K3-8 curriculum blending academic excellence with spiritual growth. Multi-age classrooms in Hartland, Delafield & Oconomowoc area."
    ),
    "/why-a-catholic-education": (
        "Why Choose Catholic Education? | SJOA Nashotah",
        "Discover the lifelong benefits of Catholic education. SJOA integrates faith, character, and academics to prepare future leaders in Waukesha County, WI."
    ),
    "/why-sjoa-school": (
        "Why SJOA? | Catholic School in Lake Country, WI",
        "Why families choose St. Joan of Arc Parish School: faith-centered education, multi-age classrooms, and a close-knit community. K3-8, Nashotah WI."
    ),
    "/schedule-a-tour": (
        "Schedule a Tour | SJOA Catholic School, Nashotah",
        "Visit St. Joan of Arc Parish School in Nashotah, WI. See our classrooms and meet our teachers. Schedule your personal tour today!"
    ),
    "/admissions": (
        "Admissions & Enrollment | K3-8 Catholic School",
        "Enroll at St. Joan of Arc - a Catholic K3-8 school in Nashotah, WI. Serving families in Hartland, Delafield, Oconomowoc & Pewaukee. Apply today."
    ),
    "/tuition-and-registration": (
        "Tuition & Registration | Catholic School, Nashotah",
        "View affordable K3-8 tuition rates, registration fees, and financial aid options at St. Joan of Arc Parish School in Nashotah, Wisconsin."
    ),
    "/contact": (
        "Contact Us | St. Joan of Arc School, Nashotah",
        "Contact St. Joan of Arc Parish School at 262-646-5821 or visit us at 120 Nashotah Road, Nashotah, WI 53058. We would love to hear from you!"
    ),
    "/3k-and-4k-catholic-preschool": (
        "3K & 4K Catholic Preschool | Hartland-Delafield",
        "Play-based K3 & K4 preschool at St. Joan of Arc in Nashotah, WI. Half-day and full-day options for families in Hartland, Delafield & Lake Country."
    ),
    "/5k-kindergarten": (
        "5K Kindergarten | Catholic School, Nashotah WI",
        "Full-day 5K Kindergarten at St. Joan of Arc Parish School. Blending academics, faith, and social development for Lake Country families."
    ),
    "/elementary-school": (
        "Catholic Elementary School (1-5) | Nashotah, WI",
        "K3 through 5th grade Catholic elementary school in Nashotah, WI. Serving Lake Country families with faith-based education and multi-age classrooms."
    ),
    "/middle-school": (
        "Catholic Middle School (6-8) | Oconomowoc Area",
        "Grades 6-8 Catholic middle school preparing students for high school success. Strong academics and character development in Nashotah, WI."
    ),
    "/multiage-learning": (
        "Multi-Age Learning at SJOA | Catholic K-8",
        "How multi-age classrooms at St. Joan of Arc Catholic School foster collaboration, leadership, and personalized learning for K3-8 students."
    ),
    "/school-faq": (
        "FAQ | St. Joan of Arc Catholic School, Nashotah",
        "Frequently asked questions about St. Joan of Arc Parish School - hours, lunch program, extended care, activities, and more."
    ),
    "/testimonials": (
        "Parent & Alumni Testimonials | SJOA Catholic",
        "Read testimonials from parents and alumni about the supportive community, strong academics, and faith-based education at St. Joan of Arc School."
    ),
    "/staff": (
        "Our Faculty & Staff | SJOA Catholic School",
        "Meet the dedicated teachers and staff at St. Joan of Arc Parish School in Nashotah, WI - committed to faith-centered K3-8 education."
    ),
    "/calendar": (
        "School Calendar | SJOA Catholic, Nashotah WI",
        "View the academic calendar at St. Joan of Arc Parish School - key dates, holidays, liturgies, and school events for the school year."
    ),
    "/student-life": (
        "Student Life | Catholic School, Lake Country WI",
        "Discover student life at St. Joan of Arc - enriching activities, liturgies, and hands-on learning that shape faith-filled future leaders."
    ),
    "/dress-code": (
        "Dress Code & Uniform Policy | SJOA Catholic",
        "Review the uniform dress code for St. Joan of Arc Parish School. Everything families need to know to prepare for the school year."
    ),
    "/school-awards": (
        "Student Awards & Recognition | SJOA Catholic",
        "Learn about the F.R.O.G. and PRO Awards at St. Joan of Arc - celebrating student effort, respect, and character in our Catholic K-8 community."
    ),
    "/extended-care": (
        "Before & After School Care | SJOA Nashotah WI",
        "Extended care at St. Joan of Arc Parish School runs 6:25 AM - 5:30 PM. Affordable weekly rates for Lake Country families."
    ),
    "/extracurriculars": (
        "Extracurriculars | Drama, Athletics & More | SJOA",
        "Drama Club, athletics, clubs, and more at St. Joan of Arc Catholic School. Enriching activities that complement academic success for K-8 students."
    ),
    "/parent-resources": (
        "Parent Resources | SJOA Catholic School Portal",
        "Essential resources for SJOA families - handbooks, forms, lunch menus, and contact info. Everything parents need in one place."
    ),
    "/fundraising-and-school-support": (
        "Support SJOA | Fundraising & SCRIP Program",
        "Support St. Joan of Arc Parish School through the annual appeal, SCRIP program, and community fundraising events. Every contribution matters."
    ),
    "/home-and-school-committee": (
        "Home & School Committee | Get Involved at SJOA",
        "Join the Home & School Committee at St. Joan of Arc. Help shape our school community and support events, activities, and faith-based initiatives."
    ),
    "/athletics": (
        "SJOA Athletics | Volleyball, Basketball & More",
        "Athletics for grades 5-8 at St. Joan of Arc Parish School - volleyball, basketball, cross country, and track. Building teamwork through Catholic values."
    ),
    "/alumni": (
        "SJOA Alumni | Stay Connected | Catholic K-8",
        "Stay connected with the St. Joan of Arc alumni community. News, events, and ways to support our mission of Catholic education."
    ),
    "/community-events": (
        "Community Events | SJOA Catholic School Calendar",
        "Family fun at St. Joan of Arc - from the Annual Carnival to Children's Liturgy of the Word. Events connecting our Catholic community."
    ),
}

ws.cell(row=1, column=5, value="Proposed Title (<55 ch.)").font = Font(bold=True)
ws.cell(row=1, column=6, value="Proposed Description (<160 ch.)").font = Font(bold=True)
ws.cell(row=1, column=7, value="Notes").font = Font(bold=True)

for row in range(2, ws.max_row + 1):
    url = ws.cell(row=row, column=2).value
    if url in proposals:
        title, desc = proposals[url]
        ws.cell(row=row, column=5, value=title).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=6, value=desc).alignment = Alignment(wrap_text=True, vertical="top")
    if url == "/elementary-school":
        cell = ws.cell(row=row, column=7, value='BUG: current title says "3K & 4K Preschool" but URL is /elementary-school - likely copy/paste error when LPI site was built')
        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 50
ws.column_dimensions["E"].width = 50
ws.column_dimensions["F"].width = 60
ws.column_dimensions["G"].width = 40

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")

print("\n=== Title length check ===")
overs_t = 0
for url, (title, desc) in proposals.items():
    if len(title) > 55:
        overs_t += 1
        print(f"WARN {len(title):3d} ch | {url} | {title}")
print(f"Titles over 55 chars: {overs_t}")

print("\n=== Description length check ===")
overs_d = 0
for url, (title, desc) in proposals.items():
    if len(desc) > 160:
        overs_d += 1
        print(f"WARN {len(desc):3d} ch | {url} | {desc}")
print(f"Descriptions over 160 chars: {overs_d}")
